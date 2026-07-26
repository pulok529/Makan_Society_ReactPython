from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

from fastapi import HTTPException, status
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.billing.service import BillingService
from app.modules.accounting.schemas import IncomeExpenseComparisonReport
from app.modules.messaging.service import MessagingService
from app.modules.reporting.repository import ReportingRepository
from app.modules.reporting.schemas import (
    ChargeRegisterRow,
    CollectionRow,
    DueMemberRow,
    ExpenseDetailRow,
    IncomeDetailRow,
    MemberRegisterRow,
    MemberInformationDetailReport,
    MemberInformationSummary,
    MemberSmsHistoryRow,
    ReceiptDetailLine,
    ReceiptDetailReport,
    ReportEnvelope,
    ReportFilter,
    ReportPageEnvelope,
    SingleMemberDueHistoryRow,
    SingleMemberPaymentHistoryRow,
    SingleMemberStatementReport,
    TotalCollectionRow,
    TotalDueRow,
)


class ReportingService:
    def __init__(self, db) -> None:
        self.db = db
        self.repository = ReportingRepository(db)
        self.template_env = Environment(
            loader=FileSystemLoader("app/modules/reporting/templates"),
            autoescape=select_autoescape(["html"]),
        )

    def _applied_filters(self, filters: ReportFilter) -> dict[str, str]:
        applied: dict[str, str] = {}

        if filters.from_date:
            applied["from_date"] = filters.from_date.isoformat()

        if filters.to_date:
            applied["to_date"] = filters.to_date.isoformat()

        if filters.member_id is not None:
            member = next(iter(self.repository.list_members(member_id=filters.member_id)), None)
            applied["member"] = f"{member.member_code} - {member.full_name}" if member else str(filters.member_id)

        if filters.category_id is not None:
            category = next((item for item in self.repository.list_categories() if item.id == filters.category_id), None)
            applied["category"] = category.name if category else str(filters.category_id)

        if filters.billing_period_id is not None:
            period = next((item for item in self.repository.list_periods() if item.id == filters.billing_period_id), None)
            applied["billing_period"] = period.period_name if period else str(filters.billing_period_id)

        if filters.plot_no and filters.plot_no.strip():
            applied["plot_no"] = filters.plot_no.strip()

        return applied

    @staticmethod
    def _month_start(value):
        return value.replace(day=1)

    def _latest_active_packages(self) -> dict[int, str | None]:
        packages = {item.id: item for item in self.repository.list_packages()}
        latest_active_package: dict[int, str | None] = {}
        for assignment in self.repository.list_member_packages():
            if assignment.is_active and assignment.package_id in packages:
                latest_active_package[assignment.member_id] = packages[assignment.package_id].name
        return latest_active_package

    def _member_collection_totals(self, member_ids: set[int]) -> dict[int, float]:
        totals: dict[int, float] = {member_id: 0.0 for member_id in member_ids}
        for invoice in self.repository.list_invoices_for_collection():
            if invoice.member_id in totals:
                totals[invoice.member_id] += float(invoice.total_receive_amount)
        return totals

    def _member_due_totals(self, member_ids: set[int]) -> dict[int, float]:
        totals: dict[int, float] = {member_id: 0.0 for member_id in member_ids}
        for charge in self.repository.list_charges():
            if charge.member_id in totals:
                totals[charge.member_id] += float(charge.due_amount)
        return totals

    def income_detail(self, filters: ReportFilter) -> ReportEnvelope:
        accounts = {item.id: item for item in self.repository.list_accounts()}
        rows = [
            IncomeDetailRow(
                income_id=entry.id,
                income_date=entry.income_date,
                account_code=accounts[entry.coa_id].code if entry.coa_id in accounts else None,
                account_name=accounts[entry.coa_id].name if entry.coa_id in accounts else None,
                amount=float(entry.amount),
                remarks=entry.remarks,
                created_at=entry.created_at,
            )
            for entry in self.repository.list_income_entries(from_date=filters.from_date, to_date=filters.to_date)
        ]
        return ReportEnvelope(
            report_type="income_detail",
            title="Income Detail Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "total_income_amount": round(sum(item.amount for item in rows), 2),
                "income_entry_count": len(rows),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump(mode="json") for row in rows],
        )

    def expense_detail(self, filters: ReportFilter) -> ReportEnvelope:
        accounts = {item.id: item for item in self.repository.list_accounts()}
        rows = [
            ExpenseDetailRow(
                expense_id=entry.id,
                expense_date=entry.expense_date,
                account_code=accounts[entry.coa_id].code if entry.coa_id in accounts else None,
                account_name=accounts[entry.coa_id].name if entry.coa_id in accounts else None,
                amount=float(entry.amount),
                remarks=entry.remarks,
                created_at=entry.created_at,
            )
            for entry in self.repository.list_expense_entries(from_date=filters.from_date, to_date=filters.to_date)
        ]
        return ReportEnvelope(
            report_type="expense_detail",
            title="Expense Detail Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "total_expense_amount": round(sum(item.amount for item in rows), 2),
                "expense_entry_count": len(rows),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump(mode="json") for row in rows],
        )

    def due_members(self, filters: ReportFilter) -> ReportEnvelope:
        categories = {item.id: item for item in self.repository.list_categories()}
        charges = self.repository.list_charges(
            member_id=filters.member_id,
            billing_period_id=filters.billing_period_id,
        )
        members = {
            member.id: member
            for member in self.repository.list_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
            )
        }

        grouped: dict[int, DueMemberRow] = {}
        for charge in charges:
            if charge.member_id not in members or float(charge.due_amount) <= 0:
                continue
            member = members[charge.member_id]
            existing = grouped.get(member.id)
            if existing is None:
                grouped[member.id] = DueMemberRow(
                    member_id=member.id,
                    member_code=member.member_code,
                    member_name=member.full_name,
                    category_name=categories[member.category_id].name if member.category_id in categories else None,
                    cell_no=member.cell_no,
                    total_charged=float(charge.net_amount),
                    total_due=float(charge.due_amount),
                    open_charge_count=1,
                )
            else:
                existing.total_charged += float(charge.net_amount)
                existing.total_due += float(charge.due_amount)
                existing.open_charge_count += 1

        rows = sorted(grouped.values(), key=lambda item: (item.member_code, item.member_name))
        return ReportEnvelope(
            report_type="due_members",
            title="Due Members Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "total_due_amount": round(sum(item.total_due for item in rows), 2),
                "member_count": len(rows),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump() for row in rows],
        )

    def electricity_collection(self, filters: ReportFilter) -> ReportEnvelope:
        total, total_bill, total_paid, rows = self.repository.electricity_collection(
            member_id=filters.member_id,
            plot_no=filters.plot_no,
            from_date=filters.from_date,
            to_date=filters.to_date,
        )
        return ReportEnvelope(
            report_type="electricity_collection",
            title="Electricity Collection Report",
            generated_at=datetime.now(UTC),
            row_count=total,
            totals={
                "total_transactions": total,
                "total_electricity_bill_amount": round(total_bill, 2),
                "total_collected_amount": round(total_paid, 2),
            },
            applied_filters=self._applied_filters(filters),
            rows=rows,
        )

    def collections(self, filters: ReportFilter) -> ReportEnvelope:
        members = {
            member.id: member
            for member in self.repository.list_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
            )
        }
        invoices = self.repository.list_invoices_for_collection(
            member_id=filters.member_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        )
        rows = [
            CollectionRow(
                member_id=invoice.member_id,
                member_code=members[invoice.member_id].member_code,
                member_name=members[invoice.member_id].full_name,
                receipt_no=invoice.invoice_no,
                payment_date=invoice.invoice_date,
                total_amount=float(invoice.total_receive_amount),
                discount_amount=float(invoice.discount_amount),
            )
            for invoice in invoices
            if invoice.member_id in members
        ]
        rows.sort(key=lambda item: (item.member_code or "", item.payment_date, item.receipt_no))
        return ReportEnvelope(
            report_type="collections",
            title="Collection Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "total_collected": round(sum(item.total_amount for item in rows), 2),
                "discount_amount": round(sum(item.discount_amount for item in rows), 2),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump() for row in rows],
        )

    def total_collection(self, filters: ReportFilter) -> ReportEnvelope:
        members = {
            member.id: member
            for member in self.repository.list_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
            )
        }
        grouped: dict[int, TotalCollectionRow] = {}
        for invoice in self.repository.list_invoices_for_collection(
            member_id=filters.member_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        ):
            if invoice.member_id is None or invoice.member_id not in members:
                continue
            member = members[invoice.member_id]
            row = grouped.get(member.id)
            if row is None:
                grouped[member.id] = TotalCollectionRow(
                    member_id=member.id,
                    member_code=member.member_code,
                    member_name=member.full_name,
                    plot_no=member.plot_no or member.member_id_text,
                    total_collection_amount=float(invoice.total_receive_amount),
                )
            else:
                row.total_collection_amount += float(invoice.total_receive_amount)
        rows = sorted(grouped.values(), key=lambda item: (item.member_code, item.member_name))
        return ReportEnvelope(
            report_type="total_collection",
            title="Total Collection Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={"total_collection_amount": sum(r.total_collection_amount for r in rows)},
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump() for row in rows],
        )

    def total_due(self, filters: ReportFilter) -> ReportEnvelope:
        members = {
            member.id: member
            for member in self.repository.list_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
            )
        }
        grouped: dict[int, TotalDueRow] = {}
        for charge in self.repository.list_charges(
            member_id=filters.member_id,
            billing_period_id=filters.billing_period_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        ):
            if charge.member_id not in members or float(charge.due_amount) <= 0:
                continue
            member = members[charge.member_id]
            row = grouped.get(member.id)
            if row is None:
                grouped[member.id] = TotalDueRow(
                    member_id=member.id,
                    member_code=member.member_code,
                    member_name=member.full_name,
                    plot_no=member.plot_no or member.member_id_text,
                    total_due_amount=float(charge.due_amount),
                )
            else:
                row.total_due_amount += float(charge.due_amount)
        rows = sorted(grouped.values(), key=lambda item: (item.member_code, item.member_name))
        return ReportEnvelope(
            report_type="total_due",
            title="Total Due Report",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "total_due_amount": round(sum(item.total_due_amount for item in rows), 2),
                "member_count": len(rows),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump() for row in rows],
        )

    def single_member_statement(self, filters: ReportFilter) -> SingleMemberStatementReport:
        if filters.member_id is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Member is required for this report")
        members = {member.id: member for member in self.repository.list_members(member_id=filters.member_id)}
        member = members.get(filters.member_id)
        if member is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")

        invoices = self.repository.list_invoices_for_collection(
            member_id=filters.member_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        )
        due_lines = BillingService(self.db).preview_member_dues(filters.member_id)
        due_history = [
            SingleMemberDueHistoryRow(
                head_name=line.head_name,
                period_display=line.period_display,
                total_bill=float(line.fee_amount),
                paid_amount=float(line.paid_amount),
                due_amount=float(line.due_amount),
            )
            for line in due_lines
            if float(line.due_amount) > 0
            and (filters.from_date is None or line.period_date is None or line.period_date >= self._month_start(filters.from_date))
            and (filters.to_date is None or line.period_date is None or line.period_date <= self._month_start(filters.to_date))
        ]
        due_history.sort(key=lambda item: (item.period_display or "", item.head_name))
        payment_history = [
            SingleMemberPaymentHistoryRow(
                receipt_no=invoice.invoice_no,
                payment_date=invoice.invoice_date,
                amount=float(invoice.total_receive_amount),
                discount_amount=float(invoice.discount_amount),
                notes=invoice.notes if hasattr(invoice, "notes") else None,
            )
            for invoice in invoices
        ]
        return SingleMemberStatementReport(
            member_id=member.id,
            member_code=member.member_code,
            member_name=member.full_name,
            plot_no=member.plot_no or member.member_id_text,
            total_bill=round(sum(item.total_bill for item in due_history), 2),
            paid_amount=round(sum(item.amount for item in payment_history), 2),
            due_amount=round(sum(item.due_amount for item in due_history), 2),
            applied_filters=self._applied_filters(filters),
            due_history=due_history,
            payment_history=payment_history,
        )

    def member_information_detail(self, filters: ReportFilter) -> MemberInformationDetailReport:
        if filters.member_id is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Member is required for this report")

        member = next(iter(self.repository.list_members(member_id=filters.member_id)), None)
        if member is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Member not found")

        categories = {item.id: item for item in self.repository.list_categories()}
        nominee = self.repository.get_member_nominee(member.id)
        invoices = self.repository.list_invoices_for_collection(
            member_id=filters.member_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        )
        payment_history = [
            SingleMemberPaymentHistoryRow(
                receipt_no=invoice.invoice_no,
                payment_date=invoice.invoice_date,
                amount=float(invoice.total_receive_amount),
                discount_amount=float(invoice.discount_amount),
                notes=invoice.notes if hasattr(invoice, "notes") else None,
            )
            for invoice in invoices
        ]
        due_lines = BillingService(self.db).preview_member_dues(filters.member_id)
        due_history = [
            SingleMemberDueHistoryRow(
                head_name=line.head_name,
                period_display=line.period_display,
                total_bill=float(line.fee_amount),
                paid_amount=float(line.paid_amount),
                due_amount=float(line.due_amount),
            )
            for line in due_lines
            if float(line.due_amount) > 0
            and (filters.from_date is None or line.period_date is None or line.period_date >= self._month_start(filters.from_date))
            and (filters.to_date is None or line.period_date is None or line.period_date <= self._month_start(filters.to_date))
        ]
        sms_history = [
            MemberSmsHistoryRow(
                created_at=item.created_at,
                recipient=item.recipient,
                template_name=item.template_name,
                message_body=item.message_body,
                status=item.status,
            )
            for item in MessagingService(self.db).list_messages()
            if item.member_id == filters.member_id
        ]

        return MemberInformationDetailReport(
            member_id=member.id,
            applied_filters=self._applied_filters(filters),
            member_info=MemberInformationSummary(
                member_code=member.member_code,
                full_name=member.full_name,
                plot_no=member.plot_no or member.member_id_text,
                plot_count=max(int(getattr(member, "plot_count", 1) or 1), 1),
                category_name=categories[member.category_id].name if member.category_id in categories else None,
                national_id=member.national_id,
                cell_no=member.cell_no,
                email=member.email,
                member_class=member.member_class,
                joined_on=member.joined_on,
                is_active=member.is_active,
                father_name=member.father_name,
                mother_name=member.mother_name,
                present_address=member.present_address,
                permanent_address=member.permanent_address,
                reference=member.reference,
                nominee_name=nominee.nominee_name if nominee is not None else None,
                nominee_cell=nominee.nominee_cell if nominee is not None else None,
                total_collection_amount=round(sum(item.amount for item in payment_history), 2),
                total_due_amount=round(sum(item.due_amount for item in due_history), 2),
            ),
            payment_history=payment_history,
            due_history=due_history,
            sms_history=sms_history,
        )

    def charge_register(self, filters: ReportFilter) -> ReportEnvelope:
        members = {
            member.id: member
            for member in self.repository.list_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
            )
        }
        periods = {period.id: period for period in self.repository.list_periods()}
        charges = self.repository.list_charges(
            member_id=filters.member_id,
            billing_period_id=filters.billing_period_id,
            from_date=filters.from_date,
            to_date=filters.to_date,
        )
        rows = [
            ChargeRegisterRow(
                charge_id=charge.id,
                created_at=charge.created_at,
                member_id=charge.member_id,
                member_code=members[charge.member_id].member_code if charge.member_id in members else "Unknown",
                member_name=members[charge.member_id].full_name if charge.member_id in members else "Unknown",
                billing_period_name=periods[charge.billing_period_id].period_name
                if charge.billing_period_id in periods
                else None,
                charge_type=charge.charge_type,
                status=charge.status,
                net_amount=float(charge.net_amount),
                due_amount=float(charge.due_amount),
            )
            for charge in charges
            if charge.member_id in members
        ]
        rows.sort(key=lambda item: (item.member_code, item.created_at, item.charge_id))
        return ReportEnvelope(
            report_type="charge_register",
            title="Charge Register",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "net_amount": round(sum(item.net_amount for item in rows), 2),
                "due_amount": round(sum(item.due_amount for item in rows), 2),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump(mode="json") for row in rows],
        )

    def member_register(self, filters: ReportFilter) -> ReportEnvelope:
        categories = {item.id: item for item in self.repository.list_categories()}
        members = self.repository.list_members(member_id=filters.member_id, category_id=filters.category_id, plot_no=filters.plot_no)
        member_ids = {member.id for member in members}
        collection_totals = self._member_collection_totals(member_ids)
        due_totals = self._member_due_totals(member_ids)
        rows = [
            MemberRegisterRow(
                member_id=member.id,
                member_code=member.member_code,
                full_name=member.full_name,
                plot_no=member.plot_no or member.member_id_text,
                plot_count=max(int(getattr(member, "plot_count", 1) or 1), 1),
                category_name=categories[member.category_id].name if member.category_id in categories else None,
                national_id=member.national_id,
                cell_no=member.cell_no,
                joined_on=member.joined_on,
                is_active=member.is_active,
                total_collection_amount=round(collection_totals.get(member.id, 0.0), 2),
                total_due_amount=round(due_totals.get(member.id, 0.0), 2),
            )
            for member in members
            if (filters.from_date is None or member.joined_on is None or member.joined_on >= filters.from_date)
            and (filters.to_date is None or member.joined_on is None or member.joined_on <= filters.to_date)
        ]
        return ReportEnvelope(
            report_type="member_summary",
            title="Total Member Summary",
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            totals={
                "member_count": len(rows),
                "active_members": sum(1 for row in rows if row.is_active),
                "total_collection_amount": round(sum(row.total_collection_amount for row in rows), 2),
                "total_due_amount": round(sum(row.total_due_amount for row in rows), 2),
            },
            applied_filters=self._applied_filters(filters),
            rows=[row.model_dump(mode="json") for row in rows],
        )

    def receipt_detail(self, receipt_id: int) -> ReceiptDetailReport:
        members = {member.id: member for member in self.repository.list_members()}
        receipt = self.repository.get_receipt(receipt_id)
        if receipt is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found")
        member = members.get(receipt.member_id) if receipt.member_id is not None else None
        lines = self.repository.list_receipt_lines(receipt_id)
        return ReceiptDetailReport(
            receipt_id=receipt.id,
            receipt_no=receipt.receipt_no,
            payment_date=receipt.payment_date,
            member_name=member.full_name if member is not None else None,
            member_code=member.member_code if member is not None else None,
            subtotal_amount=float(receipt.subtotal_amount),
            discount_amount=float(receipt.discount_amount),
            total_amount=float(receipt.total_amount),
            applied_filters={
                "receipt_no": receipt.receipt_no,
                **(
                    {"member": f"{member.member_code} - {member.full_name}"}
                    if member is not None
                    else {}
                ),
            },
            lines=[
                ReceiptDetailLine(line_type=line.line_type, amount=float(line.amount), charge_id=line.charge_id)
                for line in lines
            ],
        )

    def _style_worksheet(self, sheet, report_title: str):
        sheet.freeze_panes = 'A2'
        title_cell = sheet['A1']
        title_cell.font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
        )
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '#,##0.00'
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if row[0].row > 2:
                has_values = any(c.value is not None for c in row)
                if has_values and not any(isinstance(c.value, (int, float)) for c in row):
                    for cell in row:
                        if cell.value:
                            cell.font = Font(bold=True)
                            cell.fill = header_fill
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def render_html(self, report: ReportEnvelope) -> str:
        template = self.template_env.get_template("table_report.html")
        return template.render(report=report)

    def render_total_collection_html(self, report: ReportEnvelope) -> str:
        template = self.template_env.get_template("total_collection_report.html")
        return template.render(report=report)

    def render_paged_report_html(self, report: ReportPageEnvelope) -> str:
        template = self.template_env.get_template("table_report.html")
        return template.render(report=report)

    def render_receipt_html(self, report: ReceiptDetailReport) -> str:
        template = self.template_env.get_template("receipt_detail_report.html")
        return template.render(report=report)

    def render_member_statement_html(self, report: SingleMemberStatementReport) -> str:
        template = self.template_env.get_template("member_statement_report.html")
        return template.render(report=report)

    def render_member_information_detail_html(self, report: MemberInformationDetailReport) -> str:
        template = self.template_env.get_template("member_info_report.html")
        return template.render(report=report)

    def render_income_expense_html(self, report: IncomeExpenseComparisonReport) -> str:
        template = self.template_env.get_template("income_expense_report.html")
        return template.render(report=report)

    def render_xlsx(self, report: ReportEnvelope) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        sheet.append([report.title])
        sheet.append([])
        if report.rows:
            headers = [key.replace("_", " ").title() for key in report.rows[0].keys()]
            sheet.append(headers)
            for row in report.rows:
                sheet.append(list(row.values()))
        self._style_worksheet(sheet, report.title)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def render_total_collection_xlsx(self, report: ReportEnvelope) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Total Collection"
        sheet.append(["Total Collection Report"])
        sheet.append([])
        headers = ["Member ID", "Member Code", "Member Name", "Plot No", "Total Collection Amount"]
        sheet.append(headers)
        for row in report.rows:
            sheet.append([
                row.get("member_id"),
                row.get("member_code"),
                row.get("member_name"),
                row.get("plot_no"),
                row.get("total_collection_amount")
            ])
        row_count = len(report.rows)
        start_row = 4 # Row 1=Title, Row 2=Empty, Row 3=Header, Row 4=First Data
        end_row = start_row + row_count - 1 if row_count > 0 else start_row
        sum_formula = f"=SUM(E{start_row}:E{end_row})" if row_count > 0 else 0
        total_row = end_row + 1
        sheet.append(["", "", "", "Total Collection", sum_formula])
        
        self._style_worksheet(sheet, "Total Collection Report")
        
        # Style the total row
        from openpyxl.styles import Font, Border, Side
        bold_font = Font(bold=True)
        top_border = Border(top=Side(style='thin', color='000000'))
        
        for col in range(4, 6):
            cell = sheet.cell(row=total_row, column=col)
            cell.font = bold_font
            cell.border = top_border
            
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
        
    def render_paged_report_xlsx(self, report: ReportPageEnvelope) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = report.title[:31]
        sheet.append([report.title])
        sheet.append([f"Generated at: {report.generated_at.isoformat()}"])
        for key, value in report.applied_filters.items():
            sheet.append([key.replace('_', ' ').title(), value])
        sheet.append([])
        if report.items:
            headers = [h.replace('_', ' ').title() for h in report.items[0].keys()]
            sheet.append(headers)
            for row in report.items:
                sheet.append([row.get(k) for k in report.items[0].keys()])
        sheet.append([])
        if report.totals:
            for key, value in report.totals.items():
                sheet.append([f"Total {key.replace('_', ' ').title()}", value])
        self._style_worksheet(sheet, report.title)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def render_receipt_xlsx(self, report: ReceiptDetailReport) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Receipt Detail"
        sheet.append(["Receipt Detail Report"])
        sheet.append(["Receipt No", report.receipt_no])
        sheet.append(["Payment Date", report.payment_date.isoformat()])
        sheet.append(["Member Name", report.member_name or ""])
        sheet.append(["Member Code", report.member_code or ""])
        sheet.append([])
        sheet.append(["Line Type", "Charge ID", "Amount"])
        for line in report.lines:
            sheet.append([line.line_type.replace('_', ' ').title(), line.charge_id or "N/A", line.amount])
        sheet.append([])
        sheet.append(["Subtotal", "", report.subtotal_amount])
        sheet.append(["Discount", "", report.discount_amount])
        sheet.append(["Collected", "", report.total_amount])
        self._style_worksheet(sheet, "Receipt Detail Report")
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def render_member_statement_xlsx(self, report: SingleMemberStatementReport) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Member Statement"
        sheet.append(["Member Statement"])
        sheet.append(["Member Code", report.member_code])
        sheet.append(["Member Name", report.member_name])
        sheet.append(["Plot No", report.plot_no or ""])
        sheet.append([])
        sheet.append(["Dues (Billed)"])
        sheet.append(["Period", "Charge Type", "Total Bill", "Paid", "Due"])
        for row in report.due_history:
            period = row["period_display"] if isinstance(row, dict) else row.period_display
            head = row["head_name"] if isinstance(row, dict) else row.head_name
            bill = row["total_bill"] if isinstance(row, dict) else row.total_bill
            paid = row["paid_amount"] if isinstance(row, dict) else row.paid_amount
            due = row["due_amount"] if isinstance(row, dict) else row.due_amount
            sheet.append([period or "N/A", head, bill, paid, due])
        sheet.append([])
        sheet.append(["Payments (Received)"])
        sheet.append(["Date", "Receipt No", "Collected", "Discount", "Total Credited"])
        for row in report.payment_history:
            dt = row["payment_date"] if isinstance(row, dict) else row.payment_date
            rn = row["receipt_no"] if isinstance(row, dict) else row.receipt_no
            amt = row["amount"] if isinstance(row, dict) else row.amount
            dsc = row["discount_amount"] if isinstance(row, dict) else row.discount_amount
            sheet.append([dt.isoformat(), rn, amt, dsc, amt + dsc])
        sheet.append([])
        sheet.append(["Total Billed", "", "", report.total_bill])
        sheet.append(["Total Paid", "", "", report.paid_amount])
        sheet.append(["Total Due", "", "", report.due_amount])
        self._style_worksheet(sheet, "Member Statement")
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def render_member_information_detail_xlsx(self, report: MemberInformationDetailReport) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Member Info"
        info = report.member_info
        sheet.append(["Member Information Report"])
        sheet.append(["Member Code", info.member_code])
        sheet.append(["Full Name", info.full_name])
        sheet.append(["Plot No", info.plot_no or ""])
        sheet.append(["Category", info.category_name or ""])
        sheet.append(["Phone", info.cell_no or ""])
        sheet.append([])
        sheet.append(["Dues (Billed)"])
        sheet.append(["Period", "Charge Type", "Total Bill", "Paid", "Due"])
        for row in report.due_history:
            period = row["period_display"] if isinstance(row, dict) else row.period_display
            head = row["head_name"] if isinstance(row, dict) else row.head_name
            bill = row["total_bill"] if isinstance(row, dict) else row.total_bill
            paid = row["paid_amount"] if isinstance(row, dict) else row.paid_amount
            due = row["due_amount"] if isinstance(row, dict) else row.due_amount
            sheet.append([period or "N/A", head, bill, paid, due])
        sheet.append([])
        sheet.append(["Payments (Received)"])
        sheet.append(["Date", "Receipt No", "Collected", "Discount", "Total Credited"])
        for row in report.payment_history:
            dt = row["payment_date"] if isinstance(row, dict) else row.payment_date
            rn = row["receipt_no"] if isinstance(row, dict) else row.receipt_no
            amt = row["amount"] if isinstance(row, dict) else row.amount
            dsc = row["discount_amount"] if isinstance(row, dict) else row.discount_amount
            sheet.append([dt.isoformat(), rn, amt, dsc, amt + dsc])
        sheet.append([])
        sheet.append(["SMS History"])
        sheet.append(["Sent At", "Recipient", "Template", "Message"])
        for row in report.sms_history:
            created_at = row["created_at"] if isinstance(row, dict) else row.created_at
            recipient = row["recipient"] if isinstance(row, dict) else row.recipient
            template = row["template_name"] if isinstance(row, dict) else row.template_name
            msg = row["message_body"] if isinstance(row, dict) else row.message_body
            sheet.append([created_at.strftime('%Y-%m-%d %H:%M'), recipient, template or "N/A", msg])
        self._style_worksheet(sheet, "Member Information Report")
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def render_income_expense_xlsx(self, report: IncomeExpenseComparisonReport) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Income vs Expense"
        sheet.append(["Income vs Expense Report"])
        if report.from_date:
            sheet.append(["From Date", report.from_date.isoformat()])
        if report.to_date:
            sheet.append(["To Date", report.to_date.isoformat()])
        sheet.append([])
        sheet.append(["Income (Collections)"])
        sheet.append(["Category / COA", "Amount"])
        for row in report.income.rows:
            coa_name = row["coa_name"] if isinstance(row, dict) else row.coa_name
            amount = row["amount"] if isinstance(row, dict) else row.amount
            sheet.append([coa_name, amount])
        sheet.append(["Total Income", report.income.subtotal])
        sheet.append([])
        sheet.append(["Expenses (Payments)"])
        sheet.append(["Category / COA", "Amount"])
        for row in report.expense.rows:
            coa_name = row["coa_name"] if isinstance(row, dict) else row.coa_name
            amount = row["amount"] if isinstance(row, dict) else row.amount
            sheet.append([coa_name, amount])
        sheet.append(["Total Expense", report.expense.subtotal])
        sheet.append([])
        sheet.append(["Net Balance", report.net_amount])
        self._style_worksheet(sheet, "Income vs Expense Report")
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def paged_report(self, report_key: str, filters: ReportFilter, *, limit: int = 50, offset: int = 0) -> ReportPageEnvelope:
        safe_limit = min(max(limit, 1), 200)
        safe_offset = max(offset, 0)
        applied_filters = self._applied_filters(filters)
        generated_at = datetime.now(UTC)

        if report_key == "income-detail":
            total, total_amount, rows = self.repository.paged_income_detail(
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="income_detail",
                title="Income Detail Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_income_amount": round(total_amount, 2), "income_entry_count": total},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "expense-detail":
            total, total_amount, rows = self.repository.paged_expense_detail(
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="expense_detail",
                title="Expense Detail Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_expense_amount": round(total_amount, 2), "expense_entry_count": total},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "electricity-collection":
            total, total_bill, total_paid, rows = self.repository.paged_electricity_collection(
                member_id=filters.member_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="electricity_collection",
                title="Electricity Collection Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={
                    "total_transactions": total,
                    "total_electricity_bill_amount": round(total_bill, 2),
                    "total_collected_amount": round(total_paid, 2),
                },
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "collections":
            total, total_amount, discount_amount, rows = self.repository.paged_collections(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="collections",
                title="Collection Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_collected": round(total_amount, 2), "discount_amount": round(discount_amount, 2)},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "charges":
            total, net_amount, due_amount, rows = self.repository.paged_charge_register(
                member_id=filters.member_id,
                category_id=filters.category_id,
                billing_period_id=filters.billing_period_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="charge_register",
                title="Charge Register",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"net_amount": round(net_amount, 2), "due_amount": round(due_amount, 2)},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "due-members":
            total, total_due, rows = self.repository.paged_due_members(
                member_id=filters.member_id,
                category_id=filters.category_id,
                billing_period_id=filters.billing_period_id,
                plot_no=filters.plot_no,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="due_members",
                title="Due Members Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_due_amount": round(total_due, 2), "member_count": total},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "total-due":
            total, total_due, rows = self.repository.paged_total_due(
                member_id=filters.member_id,
                category_id=filters.category_id,
                billing_period_id=filters.billing_period_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="total_due",
                title="Total Due Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_due_amount": round(total_due, 2), "member_count": total},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "total-collection":
            total, total_collection, rows = self.repository.paged_total_collection(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="total_collection",
                title="Total Collection Report",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals={"total_collection_amount": round(total_collection, 2), "member_count": total},
                applied_filters=applied_filters,
                items=rows,
            )
        if report_key == "members":
            total, totals, rows = self.repository.paged_member_register(
                member_id=filters.member_id,
                category_id=filters.category_id,
                plot_no=filters.plot_no,
                from_date=filters.from_date,
                to_date=filters.to_date,
                limit=safe_limit,
                offset=safe_offset,
            )
            return ReportPageEnvelope(
                report_type="member_summary",
                title="Total Member Summary",
                generated_at=generated_at,
                total=total,
                limit=safe_limit,
                offset=safe_offset,
                totals=totals,
                applied_filters=applied_filters,
                items=rows,
            )
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unsupported paged report type")
